# -*- coding:utf-8 -*-
# @Time    :20222022/8/30上午10:36
# @Author  : yubq
# @Software: PyCharm
# @Software: PyCharm
'''
                   _ooOoo_
                  o8888888o
                  88" . "88
                  (| -_- |)
                  O\  =  /O
               ____/`---'\____
             .'  \\|     |//  `.
            /  \\|||  :  |||//  \
           /  _||||| -:- |||||-  \
           |   | \\\  -  /// |   |
           | \_|  ''\---/''  |   |
           \  .-\__  `-`  ___/-. /
         ___`. .'  /--.--\  `. . __
      ."" '<  `.___\_<|>_/___.'  >'"".
     | | :  `- \`.;`\ _ /`;.`/ - ` : | |
     \  \ `-.   \_ __\ /__ _/   .-` /  /
======`-.____`-.___\_____/___.-`____.-'======
                   `=---='
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
               佛祖保佑       永无BUG
            十年生死两茫茫，写程序，到天亮。
               千行代码，Bug何处藏。
            纵使上线又怎样，朝令改，夕断肠。
            
            领导每天新想法，天天改，日日忙。
               相顾无言，惟有泪千行。
            每晚灯火阑珊处，夜难寐，加班狂。
'''
import time
import datetime
import shutil
import openpyxl
import pandas as pd
import os
from flask import request,send_from_directory
from collections import Counter
from elasticsearch import Elasticsearch
from  utils.python2oracle import py2oracle
from utils.batch_processing import search_data_batch

es = Elasticsearch("")
progress_dict = {}

def clean_index(sql):
    pyor = py2oracle()
    try:
        res = pyor.search_data(sql)
        res_cnt = Counter(res)
        res_topn = [i[0][:] for i in res_cnt.most_common(10)]
    except Exception as e:
        res_topn = e
    return res_topn

def clean_guide(index_name):
    try:
        query = request.args.get('query')
        query = query.strip()
        if query == '' or query is None:
            response = ""
        if index_name == 'test1':
            body = {
                "query":{
                    "match":{
                        "last_word":query
                    }
                },
                "_source":["lastname"],
                "from": 0,
                "size": 1

            }
        else:
            body = {
                "query":{
                    "term":{
                        "first_name":{
                            "value":query
                        }
                    }
                }
            }

        res = es.search(index=index_name,body=body)
        guide_str = '|'.join([str(x['_source']['lastword']) for x in res['hits']['hits']])
        response = guide_str
    except Exception as e:
        response = " "
    return response

def clean_search(indexname,seachname,source):
    pyor = py2oracle()
    try:
        query = request.args.get("query")
        user = request.args.get('user')
        page = request.args.get('page')
        if page is None:
            page = 0
        else:
            page = (int(page)-1 if int(page)>0 and int(page) <= 2 else 2)
        if len(query) >= 20:
            query = query[:20]
        if indexname == 'test1':
            body = {
                "query":{
                    "match":{
                        seachname: query
                    }
                },
                "collapse":{
                    "field":"level_name.keyword"
                },
                "_source": source,
                "from": 0,
                "size": 20,
                "higlight":{
                    "fields":{
                        seachname:{}
                    },
                    "pre_tags": "<font color='red'>",
                    "post_tags": "</font>"
                }
            }
        else:
            body = {
                "query": {
                    "match": {
                        seachname: query
                    }
                },
                "_source": source,
                "from": 0,
                "size": 20,
                "higlight": {
                    "fields": {
                        seachname: {}
                    },
                    "pre_tags": "<font color='red'>",
                    "post_tags": "</font>"
                }
            }
        res = es.search(index=indexname,body=body)
        values = len(res['hits']['hits'])
        if values > 20:
            values = 20
        elif values == 0:
            response = ""
            return response
        start = 0
        end = 20
        test_name = res['hits']['hits'][0]['_source']['test_name']
        # 作为查询日志，存入数据库中
        sql = ""
        sql_res = pyor.insert_data(sql)
        if end > values:
            end = values
            response = res['hits']['hits'][start:end]
    except Exception as e:
        response = e
    return response

def clean_tempdownload(filename):
    try:
        file_path = os.getcwd()
        file_path = os.path.join(file_path,"downtemp")
        return send_from_directory(directory=file_path,path=filename,as_attachment=True)
    except Exception as e:
        return "下载失败"

def data_topn(data):
    columns_list = []
    for i ,j in enumerate(data):
        j_column = data[j].str.split('@',expand=True)
        j_column = j_column.stack()
        j_column.name = j
        if i != 0:
            if i == 1:
                data_t = pd.concat([j_column],axis=1)
            else:
                data_t = pd.concat([data_t,j_column])
            columns_list.append(j)
    data_t_new = data_t.reset_index(level=1,drop=True)
    data = data.drop(columns_list,axis=1).join(data_t_new)
    data = data.fillna("待补充")
    return data

def clean_upload(indexname,uppath):
    pyor = py2oracle()
    if request.method == "POST":
        f = request.files['file']
        user = request.form.get('user')
        uuid = request.form.get('uuid')
        if f.filename is None or f.filename == '':
            pass
        if user is None:
            pass
        if not os.path.exists(uppath):
            os.mkdir(uppath)
        upload_path = os.path.join(uppath,f.filename)
        f.save(upload_path)
        uptime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_nama = f.filename
        file_path = upload_path
        try:
            if file_nama[-3:] == 'csv':
                data = pd.read_csv(file_path,encoding='gbk')
            elif file_nama[-3:] == 'xls' or file_nama[-4:]:
                data = pd.read_excel(file_path)
            else:
                return False
        except Exception as e:
            return False
        data = data.drop_duplicates()
        data = data.dropna()
        data = data.reset_index(drop=True)
        if data.columns[0] != '输入信息':
            return "请下载模板"
        elif len(data['输入信息']) == 0:
            return "未检测到内容"
        # 上传文件日志
        sql = ""
        t0 = time.time()
        data_detail = []
        data_nums = data.shape[0]
        try:
            for i in range(data_nums):
                data.lic[i,['','','','']] = search_data_batch(data['输入信息'].loc[i])
                num_progress = round((i+1)*50 / data_nums,2)
                progress_dict[uuid] = num_progress
            data = data_topn(data)
            vaild_v = data[data[''] == '待补充'].shape[0]
            progress_dict[uuid] = 60
            t = time.time() - t0
            sql2 = ""
            for i in range(data.shape[0]):
                data_detail.append(('','',''))
            sql3 = "update "
            insert_res = pyor.insert_data(sql3)
            res_many = pyor.insert_manydata(sql2,data_detail)
            response = True
        except Exception as e:
            response = e
        return response

def clean_data_progress():
    uuid = request.args.get('uuid')
    try:
        response = {uuid:progress_dict[uuid]}
    except Exception as e:
        response = e
    return response

def clean_history(cleantype):
    pyor = py2oracle()
    try:
        user = request.args.get('user')
        filename = request.args.get('filename')
        loaddate = request.args.get('loaddate')
        page = request.args.get('page')
        page_num = request.args.get('page_num')
        if page_num is None:
            page_num = 10
        page_num = int(page_num)
        sql = "根据业务场景设计sql逻辑"
        history_res = pyor.search_data(sql)
        if page is None:
            page = 1
        start = (int(page)-1)*page_num
        end = int(page)*page_num
        values = len(history_res)
        if end >values:
            end = values
        history_res = history_res[start:end]
        response = history_res

    except Exception as e:
        response = e
    return response

# 覆盖原sheet页追加问价
def file_append(index_name,user,file_name,data):
    file_path =os.getcwd()
    down_date = str(datetime.date.today()).replace('-','')
    clean_type = 'download_file'
    down_path = '场景文件名'
    tempfile_name = '文件名'
    sheet_name = '清洗结果'
    file_path = os.path.join(file_path,clean_type,down_path,down_date,user)
    if not os.path.exists(file_path)
        os.mkdir(file_path)
    ori_path = os.path.join(os.getcwd(),'文件名',tempfile_name)
    file_name = file_name.replace('csv','xlsx')
    copy_path = os.path.join(file_path,file_name)
    shutil.copy(ori_path,copy_path)
    writer = pd.ExcelWriter(copy_path,engine='openpyxl')
    writer_book = openpyxl.load_workbook(writer.path)
    writer.book = writer_book
    idx = writer_book.sheetnames.index(sheet_name)
    writer_book.remove(writer_book.worksheets[idx])
    writer_book.create_sheet(sheet_name+'结果',idx)
    writer.sheets = dict((i.title,i) for i in writer_book.worksheets)
    data.to_excel(writer,sheet_name+'结果',index=False,encoding='utf8')
    writer.save()
    return file_path,file_name

def files

